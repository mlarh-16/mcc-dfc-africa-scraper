"""
scripts/consolidate.py
───────────────────────
Builds a single master Excel workbook from all MCC and DFC
output files. Includes a summary dashboard sheet.

Run from project root:
  python -m scripts.consolidate
"""

import re
import pandas as pd
from pathlib import Path
from datetime import datetime, timezone
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from scripts.config import AFRICA_CODE_TO_NAME, AFRICA_DFC_COUNTRIES, FILES, OUTPUT_DIR
from scripts.logger import get_logger

log = get_logger()

TODAY       = datetime.now(timezone.utc).strftime("%Y-%m-%d")
MASTER_PATH = OUTPUT_DIR / f"MCC_DFC_Africa_Master_{TODAY}.xlsx"

# ── Colour palette ────────────────────────────────────────────
BLUE_DARK   = "0C447C"   # header background
BLUE_MID    = "185FA5"   # MCC accent
TEAL_DARK   = "0F6E56"   # DFC accent
WHITE       = "FFFFFF"
LIGHT_BLUE  = "E6F1FB"   # MCC row fill
LIGHT_TEAL  = "E1F5EE"   # DFC row fill
LIGHT_GRAY  = "F1EFE8"   # summary row fill
AMBER       = "BA7517"   # highlight


# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────

def _header_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _header_font(bold: bool = True, color: str = WHITE) -> Font:
    return Font(bold=bold, color=color, size=11)


def _border() -> Border:
    thin = Side(style="thin", color="D3D1C7")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _style_header_row(ws, row: int, fill_color: str, font_color: str = WHITE):
    """Apply header styling to a worksheet row."""
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill    = _header_fill(fill_color)
            cell.font    = _header_font(color=font_color)
            cell.border  = _border()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_data_rows(ws, start_row: int, fill_color: str):
    """Apply alternating fill to data rows."""
    for i, row in enumerate(ws.iter_rows(min_row=start_row, max_row=ws.max_row)):
        for cell in row:
            cell.border = _border()
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if i % 2 == 0:
                cell.fill = _header_fill(fill_color)


def _auto_width(ws, min_width: int = 12, max_width: int = 50):
    """Auto-fit column widths."""
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value else 0 for cell in col),
            default=min_width
        )
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _freeze(ws, cell: str = "A2"):
    ws.freeze_panes = cell


def _display_country(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return "Unknown"
    upper = text.upper()
    return AFRICA_CODE_TO_NAME.get(upper, text)


def _country_series(df: pd.DataFrame) -> pd.Series:
    if "Country" in df.columns:
        series = df["Country"]
    else:
        series = df.get("Place of Performance Country Name", pd.Series(index=df.index, dtype="object"))
        if "Place of Performance Country Code" in df.columns:
            series = series.where(series.notna(), df["Place of Performance Country Code"])
    return series.apply(_display_country)


def _award_count(series: pd.Series) -> int:
    return series.nunique(dropna=True)


def _us_fiscal_year(values: pd.Series) -> pd.Series:
    dt = pd.to_datetime(values, errors="coerce")
    return dt.dt.year + (dt.dt.month >= 10).astype("Int64")

def _load(key: str) -> pd.DataFrame:
    """Safely load an output file, return empty df if missing."""
    path = FILES.get(key)
    if not path or not path.exists():
        log.warning(f"  File not found: {path}")
        return pd.DataFrame()
    try:
        df = pd.read_excel(path)
        log.info(f"  Loaded {key}: {len(df)} rows")
        return df
    except Exception as e:
        log.error(f"  Error loading {key}: {e}")
        return pd.DataFrame()


# ─────────────────────────────────────────────────────────────
# MCC grants by country
# ─────────────────────────────────────────────────────────────

def _mcc_grants_by_country(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "Award ID" not in df.columns:
        return pd.DataFrame()
    df = df.copy()
    df["Country"] = _country_series(df)
    by_country = (
        df.groupby("Country", dropna=False)
        .agg(
            Awards=("Award ID", _award_count),
            Total_Committed=("Award Amount", "sum"),
        )
        .sort_values("Total_Committed", ascending=False)
        .reset_index()
    )
    by_country.columns = ["Country", "Number of Awards", "Total Committed (USD)"]
    by_country["Total Committed (USD millions)"] = by_country["Total Committed (USD)"].apply(
        lambda x: round(x / 1_000_000, 2) if pd.notna(x) else 0
    )
    return by_country


def _year_from_award_id(aid: str) -> int | None:
    """Extract fiscal year from MCC Award ID patterns:
      - COM08TZA…  / LOG18SE2…  / GR18…  → 2–3 letters + 2-digit year
      - 95332418T…                        → digits with year at positions 5-6
    """
    s = str(aid).strip()
    # Pattern 1: 2–3 letter prefix then 2-digit year (most MCC compacts and grants)
    m = re.match(r'^[A-Za-z]{2,3}(\d{2})', s)
    if m:
        yy = int(m.group(1))
        if 4 <= yy <= 30:   # MCC first compact FY2004; cap at FY2030
            return 2000 + yy
    # Pattern 2: 953324{YY}T… (USASpending administrative grant contracts)
    m2 = re.match(r'^953324(\d{2})T', s)
    if m2:
        yy = int(m2.group(1))
        if 4 <= yy <= 30:
            return 2000 + yy
    return None


def _mcc_grants_by_year(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "Award ID" not in df.columns:
        return pd.DataFrame()
    df = df.copy()
    # USASpending does not populate Start Date for financial assistance grants.
    # Cascade per-row: Fiscal Year → Start Date → End Date → Award ID.
    fiscal_year = df.get("Fiscal Year", pd.Series(index=df.index, dtype="object"))
    if "Start Date" in df.columns:
        fiscal_year = fiscal_year.where(fiscal_year.notna(), _us_fiscal_year(df["Start Date"]))
    if "End Date" in df.columns:
        fiscal_year = fiscal_year.where(fiscal_year.notna(), _us_fiscal_year(df["End Date"]))
    fiscal_year = fiscal_year.where(fiscal_year.notna(), df["Award ID"].apply(_year_from_award_id))
    # Cast to object so downstream can mix ints and 'Not Specified'
    df["Fiscal Year"] = fiscal_year.astype(object)

    _year_known = df["Fiscal Year"].notna() & (df["Fiscal Year"].astype(str) != "Not Specified")
    df_known   = df[_year_known].copy()
    df_unknown = df[~_year_known].copy()

    parts = []

    if not df_known.empty:
        df_known["Fiscal Year"] = df_known["Fiscal Year"].astype(int)
        by_year = (
            df_known.groupby("Fiscal Year")
            .agg(
                Awards=("Award ID", _award_count),
                Total_Committed=("Award Amount", "sum"),
            )
            .sort_values("Fiscal Year")
            .reset_index()
        )
        by_year.columns = ["Fiscal Year", "Number of Awards", "Total Committed (USD)"]
        by_year["Total Committed (USD millions)"] = by_year["Total Committed (USD)"].apply(
            lambda x: round(x / 1_000_000, 2) if pd.notna(x) else 0
        )
        parts.append(by_year)

    if not df_unknown.empty:
        unknown_amt = df_unknown["Award Amount"].sum()
        parts.append(pd.DataFrame([{
            "Fiscal Year":                  "Not Specified",
            "Number of Awards":             _award_count(df_unknown["Award ID"]),
            "Total Committed (USD)":        unknown_amt,
            "Total Committed (USD millions)": round(unknown_amt / 1_000_000, 2),
        }]))

    if not parts:
        return pd.DataFrame()

    return pd.concat(parts, ignore_index=True)


def _fill_grants_fiscal_year(df: pd.DataFrame) -> pd.DataFrame:
    """Fill Fiscal Year on the raw grants sheet, cascading per-row through
    all available sources: Fiscal Year → Start Date → End Date → Award ID.
    Rows with no resolvable year get 'Not Specified'."""
    if df.empty or "Award ID" not in df.columns:
        return df
    df = df.copy()
    fy = df.get("Fiscal Year", pd.Series(index=df.index, dtype="object"))
    if "Start Date" in df.columns:
        fy = fy.where(fy.notna(), _us_fiscal_year(df["Start Date"]))
    if "End Date" in df.columns:
        fy = fy.where(fy.notna(), _us_fiscal_year(df["End Date"]))
    fy = fy.where(fy.notna(), df["Award ID"].apply(_year_from_award_id))
    # Cast to object so we can mix ints and the 'Not Specified' string
    fy_obj = fy.astype(object).where(fy.notna(), "Not Specified")
    df["Fiscal Year"] = fy_obj
    return df


# MCC year-by-year analysis
# ─────────────────────────────────────────────────────────────

def _mcc_by_year(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "Award ID" not in df.columns:
        return pd.DataFrame()
    df = df.copy()
    if "Fiscal Year" in df.columns:
        fiscal_year = df["Fiscal Year"]
    else:
        fiscal_year = _us_fiscal_year(df.get("Start Date"))
        end_fy = _us_fiscal_year(df.get("End Date"))
        fiscal_year = fiscal_year.where(fiscal_year.notna(), end_fy)
    df["Fiscal Year"] = fiscal_year
    df = df[df["Fiscal Year"].notna()].copy()
    df["Fiscal Year"] = df["Fiscal Year"].astype(int)
    by_year = (
        df.groupby("Fiscal Year")
        .agg(
            Awards=("Award ID", _award_count),
            Total_USD=("Award Amount", "sum"),
        )
        .reset_index()
        .sort_values("Fiscal Year")
    )
    by_year["Total_USD_M"] = by_year["Total_USD"].apply(
        lambda x: round(x / 1_000_000, 2) if pd.notna(x) else 0
    )
    by_year.rename(columns={
        "Awards": "Number of Awards",
        "Total_USD": "Total Award (USD)",
        "Total_USD_M": "Total Award (USD millions)",
    }, inplace=True)
    return by_year


def _mcc_top_recipients_africa(df: pd.DataFrame, n: int = 20) -> pd.DataFrame:
    if df.empty or "Recipient Name" not in df.columns or "Award Amount" not in df.columns:
        return pd.DataFrame()
    top = (
        df.groupby("Recipient Name")["Award Amount"]
        .sum()
        .sort_values(ascending=False)
        .head(n)
        .reset_index()
    )
    top.columns = ["Recipient Name", "Total Award (USD)"]
    top["Total Award (USD millions)"] = top["Total Award (USD)"].apply(
        lambda x: round(x / 1_000_000, 2)
    )
    top["Rank"] = range(1, len(top) + 1)
    return top[["Rank", "Recipient Name", "Total Award (USD)", "Total Award (USD millions)"]]


def _mcc_by_country(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "Award ID" not in df.columns:
        return pd.DataFrame()
    df = df.copy()
    df["Country"] = _country_series(df)
    by_country = (
        df.groupby("Country", dropna=False)
        .agg(
            Awards=("Award ID", _award_count),
            Total_USD=("Award Amount", "sum"),
        )
        .sort_values("Total_USD", ascending=False)
        .reset_index()
    )
    by_country.columns = [
        "Country", "Number of Awards", "Total Award (USD)"
    ]
    by_country["Total Award (USD millions)"] = by_country["Total Award (USD)"].apply(
        lambda x: round(x / 1_000_000, 2) if pd.notna(x) else 0
    )
    return by_country


# ─────────────────────────────────────────────────────────────
# DFC year-by-year analysis
# ─────────────────────────────────────────────────────────────

def _dfc_by_year(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "Fiscal Year" not in df.columns:
        return pd.DataFrame()
    by_year = (
        df.groupby("Fiscal Year")
        .agg(
            Projects=("Project Name", "count"),
            Total_Committed=("Committed", "sum"),
        )
        .reset_index()
        .sort_values("Fiscal Year")
    )
    by_year["Total_Committed_M"] = by_year["Total_Committed"].apply(
        lambda x: round(x / 1_000_000, 2) if pd.notna(x) else 0
    )
    by_year.rename(columns={
        "Projects":          "Number of Projects",
        "Total_Committed":   "Total Committed (USD)",
        "Total_Committed_M": "Total Committed (USD millions)",
    }, inplace=True)
    return by_year


_SPECIFIC_COUNTRIES = {c.lower() for c in AFRICA_DFC_COUNTRIES}


def _is_specific_country(country: str) -> bool:
    """True if any semicolon-separated part is a known specific country."""
    parts = [p.strip().lower() for p in str(country).split(";")]
    return any(p in _SPECIFIC_COUNTRIES for p in parts)


def _dfc_by_country(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "Country" not in df.columns:
        return pd.DataFrame()
    df_specific = df[df["Country"].apply(_is_specific_country)].copy()
    if df_specific.empty:
        return pd.DataFrame()

    df_specific["Country"] = df_specific["Country"].astype(str).str.split(r"\s*;\s*")
    df_exploded = df_specific.explode("Country")
    df_exploded["Country"] = df_exploded["Country"].str.strip()
    df_exploded = df_exploded[df_exploded["Country"].str.lower().isin(_SPECIFIC_COUNTRIES)].copy()
    if df_exploded.empty:
        return pd.DataFrame()

    split_counts = df_specific["Country"].apply(lambda x: len([p for p in x if p and p.strip().lower() in _SPECIFIC_COUNTRIES]))
    split_count_map = split_counts.to_dict()
    df_exploded["split_count"] = df_exploded.index.map(split_count_map).fillna(1).astype(float)
    df_exploded["Allocated Committed"] = pd.to_numeric(df_exploded.get("Committed"), errors="coerce") / df_exploded["split_count"]

    by_country = (
        df_exploded.groupby("Country")
        .agg(
            Projects=("Project Name", "count"),
            Total_Committed=("Allocated Committed", "sum"),
        )
        .sort_values("Total_Committed", ascending=False)
        .reset_index()
    )
    by_country.columns = [
        "Country", "Number of Project-Country Mentions", "Allocated Committed (USD)"
    ]
    by_country["Allocated Committed (USD millions)"] = by_country["Allocated Committed (USD)"].apply(
        lambda x: round(x / 1_000_000, 2) if pd.notna(x) else 0
    )
    return by_country


def _dfc_regional_projects(df: pd.DataFrame) -> pd.DataFrame:
    """Rows where Country is a regional label (e.g. 'Sub-Saharan Africa')
    rather than a specific country. Kept separate to avoid double-counting."""
    if df.empty or "Country" not in df.columns:
        return pd.DataFrame()
    return df[~df["Country"].apply(_is_specific_country)].copy()


def _dfc_worldwide_projects(df: pd.DataFrame) -> pd.DataFrame:
    """DFC projects tagged as 'Worldwide' region — excluded from the Africa
    total but included here for reference. Some (e.g. Gavi, SSA-labelled funds)
    have material Africa exposure. NOTE: these are global commitments; the full
    amount should not be added to Africa totals."""
    if df.empty or "Region" not in df.columns:
        return pd.DataFrame()
    df_ww = df[df["Region"].str.strip().str.lower() == "worldwide"].copy()
    if "Committed" in df_ww.columns:
        df_ww["Committed"] = pd.to_numeric(df_ww["Committed"], errors="coerce")
        df_ww["Committed (USD millions)"] = df_ww["Committed"].apply(
            lambda x: round(x / 1_000_000, 2) if pd.notna(x) else 0
        )
    return df_ww


def _dfc_by_sector(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "NAICS Sector" not in df.columns:
        return pd.DataFrame()
    by_sector = (
        df.groupby("NAICS Sector")
        .agg(
            Projects=("Project Name", "count"),
            Total_Committed=("Committed", "sum"),
        )
        .sort_values("Total_Committed", ascending=False)
        .reset_index()
    )
    by_sector.columns = [
        "NAICS Sector", "Number of Projects", "Total Committed (USD)"
    ]
    by_sector["Total Committed (USD millions)"] = by_sector["Total Committed (USD)"].apply(
        lambda x: round(x / 1_000_000, 2) if pd.notna(x) else 0
    )
    return by_sector


def _mcc_compact_sectors_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Summarise compact grant spending by high-level sector.
    Rows sourced from a budget table are summed directly.
    Rows sourced from a compact total (no breakdown) are listed separately."""
    if df.empty or "Sector" not in df.columns:
        return pd.DataFrame()
    df = df.copy()
    df["Amount (USD)"] = pd.to_numeric(df["Amount (USD)"], errors="coerce").fillna(0)

    # Exclude admin/M&E from the sector totals (keep in raw sheet only)
    exclude = {"Monitoring & Evaluation", "Program Administration"}
    df_prog = df[~df["Sector"].isin(exclude)].copy()

    # Rows with a proper breakdown vs. compact-total-only rows
    df_table = df_prog[df_prog["Source"] == "Budget Table"]
    df_total = df_prog[df_prog["Source"] != "Budget Table"]

    parts = []
    if not df_table.empty:
        by_sector = (
            df_table.groupby("Sector")["Amount (USD)"]
            .sum()
            .sort_values(ascending=False)
            .reset_index()
        )
        by_sector.columns = ["Sector", "Total (USD)"]
        by_sector["Total (USD millions)"] = (by_sector["Total (USD)"] / 1_000_000).round(2)
        by_sector["Source"] = "Component breakdown"
        parts.append(by_sector)

    if not df_total.empty:
        by_sector_t = (
            df_total.groupby("Sector")["Amount (USD)"]
            .sum()
            .sort_values(ascending=False)
            .reset_index()
        )
        by_sector_t.columns = ["Sector", "Total (USD)"]
        by_sector_t["Total (USD millions)"] = (by_sector_t["Total (USD)"] / 1_000_000).round(2)
        by_sector_t["Source"] = "Compact total only (no breakdown)"
        parts.append(by_sector_t)

    if not parts:
        return pd.DataFrame()
    return pd.concat(parts, ignore_index=True)


def _mcc_by_sector(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "naics_description" not in df.columns:
        return pd.DataFrame()
    df = df.copy()
    df["naics_description"] = df["naics_description"].fillna("Not Specified")
    by_sector = (
        df.groupby("naics_description")
        .agg(
            Awards=("Award ID", _award_count),
            Total_USD=("Award Amount", "sum"),
        )
        .sort_values("Total_USD", ascending=False)
        .reset_index()
    )
    by_sector.columns = ["NAICS Sector", "Number of Awards", "Total Award (USD)"]
    by_sector["Total Award (USD millions)"] = by_sector["Total Award (USD)"].apply(
        lambda x: round(x / 1_000_000, 2) if pd.notna(x) else 0
    )
    return by_sector


def _mcc_by_recipient_country(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "recipient_location_country_name" not in df.columns:
        return pd.DataFrame()
    df = df.copy()
    df["recipient_location_country_name"] = df["recipient_location_country_name"].fillna("Not Specified")
    by_country = (
        df.groupby("recipient_location_country_name")
        .agg(
            Awards=("Award ID", _award_count),
            Total_USD=("Award Amount", "sum"),
        )
        .sort_values("Total_USD", ascending=False)
        .reset_index()
    )
    by_country.columns = ["Recipient Country", "Number of Awards", "Total Award (USD)"]
    by_country["Total Award (USD millions)"] = by_country["Total Award (USD)"].apply(
        lambda x: round(x / 1_000_000, 2) if pd.notna(x) else 0
    )
    return by_country


def _mcc_recipients_africa(df_awards: pd.DataFrame, df_grants: pd.DataFrame) -> pd.DataFrame:
    frames = []
    for df in [df_awards, df_grants]:
        if not df.empty and {"Recipient Name", "Award Amount"}.issubset(df.columns):
            frames.append(df[["Recipient Name", "Award Amount"]].copy())
    if not frames:
        return pd.DataFrame()
    out = (
        pd.concat(frames, ignore_index=True)
        .groupby("Recipient Name", dropna=False)["Award Amount"]
        .sum()
        .sort_values(ascending=False)
        .reset_index()
    )
    out.columns = ["Recipient Name", "Total Africa Awards and Grants (USD)"]
    out["Total Africa Awards and Grants (USD millions)"] = (out["Total Africa Awards and Grants (USD)"] / 1_000_000).round(2)
    return out


def _build_qa_checks(df_mcc_awards: pd.DataFrame, df_mcc_awards_year: pd.DataFrame, df_mcc_awards_country: pd.DataFrame,
                     df_mcc_grants: pd.DataFrame, df_mcc_grants_year: pd.DataFrame, df_mcc_grants_country: pd.DataFrame,
                     df_dfc_projects: pd.DataFrame, df_dfc_by_country: pd.DataFrame) -> pd.DataFrame:
    rows = []

    def add_check(category: str, metric: str, raw_value, summary_value, notes: str = ""):
        diff = None
        try:
            if pd.notna(raw_value) and pd.notna(summary_value):
                diff = float(raw_value) - float(summary_value)
        except Exception:
            diff = None
        rows.append({
            "Category": category,
            "Check": metric,
            "Raw Value": raw_value,
            "Summary Value": summary_value,
            "Difference": diff,
            "Notes": notes,
        })

    if not df_mcc_awards.empty:
        add_check("MCC Awards", "Raw total vs by-year total", pd.to_numeric(df_mcc_awards.get("Award Amount"), errors="coerce").sum(), pd.to_numeric(df_mcc_awards_year.get("Total Award (USD)"), errors="coerce").sum(), "Should reconcile except for rows with missing year if excluded.")
        add_check("MCC Awards", "Raw total vs by-country total", pd.to_numeric(df_mcc_awards.get("Award Amount"), errors="coerce").sum(), pd.to_numeric(df_mcc_awards_country.get("Total Award (USD)"), errors="coerce").sum(), "Should reconcile when all rows have a country or fallback code.")
        add_check("MCC Awards", "Rows missing Fiscal Year", int(df_mcc_awards["Fiscal Year"].isna().sum()) if "Fiscal Year" in df_mcc_awards.columns else None, 0, "Investigate if non-zero.")

    if not df_mcc_grants.empty:
        add_check("MCC Grants", "Raw total vs by-year total", pd.to_numeric(df_mcc_grants.get("Award Amount"), errors="coerce").sum(), pd.to_numeric(df_mcc_grants_year.get("Total Committed (USD)"), errors="coerce").sum(), "Difference usually equals grants with unknown Fiscal Year.")
        add_check("MCC Grants", "Raw total vs by-country total", pd.to_numeric(df_mcc_grants.get("Award Amount"), errors="coerce").sum(), pd.to_numeric(df_mcc_grants_country.get("Total Committed (USD)"), errors="coerce").sum(), "Should reconcile when country fallback works.")
        _fy_missing = df_mcc_grants["Fiscal Year"].isna() | (df_mcc_grants["Fiscal Year"].astype(str) == "Not Specified") if "Fiscal Year" in df_mcc_grants.columns else pd.Series(False, index=df_mcc_grants.index)
        unknown_amt = pd.to_numeric(df_mcc_grants.loc[_fy_missing, "Award Amount"], errors="coerce").sum()
        add_check("MCC Grants", "Amount with missing Fiscal Year", unknown_amt, 0, "Large values indicate date gaps in source data.")

    if not df_dfc_projects.empty and not df_dfc_by_country.empty:
        add_check("DFC Projects", "Specific-country total vs allocated by-country total", pd.to_numeric(df_dfc_projects.loc[df_dfc_projects["Country"].apply(_is_specific_country), "Committed"], errors="coerce").sum(), pd.to_numeric(df_dfc_by_country.get("Allocated Committed (USD)"), errors="coerce").sum(), "Should reconcile after equal allocation across multi-country rows.")

    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────
# Summary sheet builder
# ─────────────────────────────────────────────────────────────

def _build_summary(
    df_mcc: pd.DataFrame,
    df_dfc: pd.DataFrame,
    df_recipients: pd.DataFrame,
    df_sectors: pd.DataFrame,
) -> pd.DataFrame:
    """Build a high-level summary table."""
    rows = []

    # MCC stats
    if not df_mcc.empty:
        rows.append({
            "Category":    "MCC Africa",
            "Metric":      "Total Africa contract awards (USASpending)",
            "Value":       len(df_mcc),
            "USD":         df_mcc["Award Amount"].sum() if "Award Amount" in df_mcc.columns else 0,
            "Notes":       (lambda s, e: f"Start dates: {s} to {e}" if pd.notna(s) and pd.notna(e) else "")(df_mcc["Start Date"].min(), df_mcc["Start Date"].max()) if "Start Date" in df_mcc.columns else "",
        })

    if not df_recipients.empty:
        rows.append({
            "Category": "MCC Africa",
            "Metric":   "Top recipient firms tracked",
            "Value":    len(df_recipients),
            "USD":      0,
            "Notes":    "Recipient-category rows from the MCC contract-award pull",
        })

    # DFC stats
    if not df_dfc.empty:
        rows.append({
            "Category": "DFC Africa",
            "Metric":   "Total projects (FY2024 database)",
            "Value":    len(df_dfc),
            "USD":      df_dfc["Committed"].sum() if "Committed" in df_dfc.columns else 0,
            "Notes":    "Source: DFC Annual Project Data FY2024",
        })

    if not df_sectors.empty:
        rows.append({
            "Category": "DFC Africa",
            "Metric":   "Sector story pages scraped",
            "Value":    len(df_sectors),
            "USD":      0,
            "Notes":    "Energy, Agri, Health, Infrastructure, Finance",
        })

    df_summary = pd.DataFrame(rows)
    if not df_summary.empty:
        df_summary["USD (millions)"] = df_summary["USD"].apply(
            lambda x: round(x / 1_000_000, 2) if pd.notna(x) and x > 0 else ""
        )
    return df_summary


# ─────────────────────────────────────────────────────────────
# Write master workbook
# ─────────────────────────────────────────────────────────────

def build_master_workbook():
    log.info("=" * 55)
    log.info("  Building master Excel workbook")
    log.info("=" * 55)

    # ── Load all source files ─────────────────────────────────
    df_mcc_africa      = _load("usaspending_africa")
    df_mcc_recipients  = _load("usaspending_recipients")
    df_mcc_countries   = _load("mcc_countries")
    df_mcc_grants          = _fill_grants_fiscal_year(_load("mcc_grants_africa"))
    df_mcc_compact_sectors = _load("mcc_compact_sectors")
    df_mcc_open_data       = _load("mcc_open_data_sectors")
    df_dfc_projects    = _load("dfc_active_projects")
    df_dfc_raw         = _load("dfc_raw_download")
    df_dfc_stories     = _load("dfc_impact_stories")
    df_dfc_press       = _load("dfc_press_releases")
    df_dfc_fr          = _load("dfc_federal_register")
    df_dfc_spending    = _load("dfc_usaspending")
    df_dfc_sectors     = _load("dfc_sectors")

    # ── Derived tables ────────────────────────────────────────
    df_mcc_by_year         = _mcc_by_year(df_mcc_africa)
    df_mcc_top             = _mcc_top_recipients_africa(df_mcc_africa)
    df_mcc_by_country      = _mcc_by_country(df_mcc_africa)
    df_mcc_by_sector            = _mcc_by_sector(df_mcc_africa)
    df_mcc_by_recip_country     = _mcc_by_recipient_country(df_mcc_africa)
    df_mcc_grants_country  = _mcc_grants_by_country(df_mcc_grants)
    df_mcc_grants_year     = _mcc_grants_by_year(df_mcc_grants)
    df_mcc_recipients_africa = _mcc_recipients_africa(df_mcc_africa, df_mcc_grants)
    df_dfc_by_year     = _dfc_by_year(df_dfc_projects)
    df_dfc_by_country  = _dfc_by_country(df_dfc_projects)
    df_dfc_regional    = _dfc_regional_projects(df_dfc_projects)
    df_dfc_worldwide   = _dfc_worldwide_projects(df_dfc_raw)
    df_dfc_by_sector   = _dfc_by_sector(df_dfc_projects)
    df_summary         = _build_summary(
        df_mcc_africa, df_dfc_projects,
        df_mcc_recipients, df_dfc_sectors
    )
    df_qa_checks = _build_qa_checks(
        df_mcc_africa, df_mcc_by_year, df_mcc_by_country,
        df_mcc_grants, df_mcc_grants_year, df_mcc_grants_country,
        df_dfc_projects, df_dfc_by_country
    )

    # ── Sheet definitions ─────────────────────────────────────
    # (sheet_name, dataframe, header_color, row_fill_color)
    sheets = [
        ("00 Summary",                   df_summary,          BLUE_DARK,  LIGHT_GRAY),
        ("00 QA Checks",                 df_qa_checks,        AMBER,      LIGHT_GRAY),
        ("MCC — Awards by Year",          df_mcc_by_year,          BLUE_MID,   LIGHT_BLUE),
        ("MCC — Awards by Country",       df_mcc_by_country,       BLUE_MID,   LIGHT_BLUE),
        ("MCC — Awards by Sector",        df_mcc_by_sector,        BLUE_MID,   LIGHT_BLUE),
        ("MCC — Awards by Firm Country",  df_mcc_by_recip_country, BLUE_MID,   LIGHT_BLUE),
        ("MCC — Top Recipients (Africa)", df_mcc_top,              BLUE_MID,   LIGHT_BLUE),
        ("MCC — All Africa Awards",       df_mcc_africa,       BLUE_MID,   LIGHT_BLUE),
        ("MCC — Recipients (Africa)",     df_mcc_recipients_africa, BLUE_MID, LIGHT_BLUE),
        ("MCC — Global Contracts",         df_mcc_recipients,   BLUE_MID,   LIGHT_BLUE),
        ("MCC — Country Page Summaries",  df_mcc_countries,    BLUE_MID,   LIGHT_BLUE),
        ("MCC — Grants by Country",       df_mcc_grants_country,       BLUE_MID, LIGHT_BLUE),
        ("MCC — Grants by Year",          df_mcc_grants_year,          BLUE_MID, LIGHT_BLUE),
        ("MCC — All Africa Grants",       df_mcc_grants,               BLUE_MID, LIGHT_BLUE),
        ("MCC — Sectors (MCC.gov)",       df_mcc_compact_sectors,      BLUE_MID, LIGHT_BLUE),
        ("MCC — Sectors (Open Data)",     df_mcc_open_data,            BLUE_MID, LIGHT_BLUE),
        ("DFC — Projects by Year",        df_dfc_by_year,      TEAL_DARK,  LIGHT_TEAL),
        ("DFC — Projects by Country",     df_dfc_by_country,   TEAL_DARK,  LIGHT_TEAL),
        ("DFC — Regional Projects",       df_dfc_regional,     TEAL_DARK,  LIGHT_TEAL),
        ("DFC — Worldwide Projects",      df_dfc_worldwide,    TEAL_DARK,  LIGHT_TEAL),
        ("DFC — Projects by Sector",      df_dfc_by_sector,    TEAL_DARK,  LIGHT_TEAL),
        ("DFC — All Africa Projects",     df_dfc_projects,     TEAL_DARK,  LIGHT_TEAL),
        ("DFC — Investment Stories",      df_dfc_stories,      TEAL_DARK,  LIGHT_TEAL),
        ("DFC — Press Releases",          df_dfc_press,        TEAL_DARK,  LIGHT_TEAL),
        ("DFC — Federal Register",         df_dfc_fr,           TEAL_DARK,  LIGHT_TEAL),
        ("DFC — Africa Contracts",         df_dfc_spending,     TEAL_DARK,  LIGHT_TEAL),
        ("DFC — Sector Stories",          df_dfc_sectors,      TEAL_DARK,  LIGHT_TEAL),
    ]

    # ── Write to Excel ────────────────────────────────────────
    with pd.ExcelWriter(MASTER_PATH, engine="openpyxl") as writer:
        for sheet_name, df, header_color, row_color in sheets:
            if df.empty:
                log.warning(f"  Skipping empty sheet: {sheet_name}")
                continue
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            log.info(f"  Written: '{sheet_name}' ({len(df)} rows)")

    # ── Apply styling ─────────────────────────────────────────
    log.info("  Applying styles...")
    wb = load_workbook(MASTER_PATH)

    for sheet_name, df, header_color, row_color in sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        _style_header_row(ws, row=1, fill_color=header_color)
        _style_data_rows(ws, start_row=2, fill_color=row_color)
        _auto_width(ws)
        _freeze(ws)
        ws.row_dimensions[1].height = 30

    wb.save(MASTER_PATH)
    log.info(f"\n  Master workbook saved -> {MASTER_PATH}")
    written = [name for name, df, _, _ in sheets if not df.empty]
    log.info(f"  Sheets written: {len(written)}")


# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────

def main():
    build_master_workbook()
    log.info("\n  Done! Open the master workbook in Excel.")
    log.info(f"  File: {MASTER_PATH.name}")


if __name__ == "__main__":
    main()
